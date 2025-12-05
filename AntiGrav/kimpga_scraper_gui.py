"""
김프가(kimpga.com) 상위 코인 크롤링 스크립트 - PyQt5 GUI 버전
BeautifulSoup과 Selenium을 함께 사용하여 더 안정적으로 데이터를 추출합니다.
"""

import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
                             QLabel, QSpinBox, QProgressBar, QTextEdit, QFileDialog,
                             QGroupBox, QCheckBox, QMessageBox, QHeaderView)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QFont, QColor, QIcon

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
import time
from datetime import datetime
import re


class ScraperThread(QThread):
    """백그라운드에서 스크래핑을 수행하는 스레드"""
    
    # 시그널 정의
    progress_update = pyqtSignal(str)  # 진행 상황 메시지
    data_scraped = pyqtSignal(list)    # 스크래핑된 데이터
    finished = pyqtSignal()            # 작업 완료
    error = pyqtSignal(str)            # 에러 발생
    
    def __init__(self, num_coins=20, headless=True):
        super().__init__()
        self.num_coins = num_coins
        self.headless = headless
        self.url = "https://kimpga.com/"
        self.driver = None
        self.is_running = True
        
    def setup_driver(self):
        """Chrome WebDriver 설정"""
        chrome_options = Options()
        
        if self.headless:
            chrome_options.add_argument('--headless=new')
        
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
    def stop(self):
        """스크래핑 중지"""
        self.is_running = False
        if self.driver:
            self.driver.quit()
    
    def run(self):
        """스레드 실행"""
        try:
            self.setup_driver()
            self.progress_update.emit(f"[WEB] {self.url} 접속 중...")
            
            # 페이지 로드
            self.driver.get(self.url)
            
            if not self.is_running:
                return
            
            # 페이지가 완전히 로드될 때까지 대기
            self.progress_update.emit("[WAIT] 페이지 로딩 대기 중...")
            time.sleep(5)
            
            if not self.is_running:
                return
            
            # 스크롤하여 모든 데이터 로드
            self.progress_update.emit("[SCROLL] 페이지 스크롤 중...")
            for _ in range(3):
                if not self.is_running:
                    return
                self.driver.execute_script("window.scrollBy(0, 500);")
                time.sleep(0.5)
            
            # 페이지 소스 가져오기
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # 코인 데이터 추출
            coins_data = []
            
            self.progress_update.emit(f"[DATA] 상위 {self.num_coins}개 코인 데이터 추출 중...")
            
            # 다양한 방법으로 코인 행 찾기
            rows = soup.find_all('div', {'role': 'row'})
            
            if not rows:
                rows = soup.find_all('tr')
            
            if not rows:
                self.error.emit("코인 데이터를 찾을 수 없습니다.")
                return
            
            self.progress_update.emit(f"[OK] {len(rows)}개의 행 발견")
            
            count = 0
            for idx, row in enumerate(rows):
                if not self.is_running:
                    return
                    
                if count >= self.num_coins:
                    break
                
                try:
                    # 이미지 태그가 있는지 확인 (코인 아이콘)
                    img = row.find('img')
                    if not img:
                        continue
                    
                    # 모든 텍스트 추출
                    row_text = row.get_text(separator='|', strip=True)
                    text_parts = [part.strip() for part in row_text.split('|') if part.strip()]
                    
                    # 헤더 행 건너뛰기
                    if any(keyword in row_text for keyword in ['순위', '코인명', 'Rank', 'Name']):
                        continue
                    
                    # 코인 정보 추출
                    coin_name = ""
                    coin_symbol = ""
                    
                    # span 태그에서 코인명과 심볼 찾기
                    spans = row.find_all('span')
                    for span in spans:
                        span_text = span.get_text(strip=True)
                        if span_text and len(span_text) > 1:
                            if not coin_name and not any(char.isdigit() for char in span_text[:3]):
                                coin_name = span_text
                            elif not coin_symbol and span_text != coin_name and not any(char.isdigit() for char in span_text):
                                coin_symbol = span_text
                                break
                    
                    # 코인명이 없으면 건너뛰기
                    if not coin_name:
                        continue
                    
                    # 기본 데이터 구조
                    coin_data = {
                        '순위': count + 1,
                        '코인명': coin_name,
                        '심볼': coin_symbol if coin_symbol else coin_name,
                        '수집시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # 가격 및 프리미엄 정보 추출
                    price_info = []
                    for part in text_parts:
                        if re.search(r'\d', part):
                            price_info.append(part)
                    
                    # 가격 정보 할당
                    if len(price_info) >= 1:
                        coin_data['국내가격'] = price_info[0]
                    if len(price_info) >= 2:
                        coin_data['해외가격'] = price_info[1]
                    if len(price_info) >= 3:
                        coin_data['김프율'] = price_info[2]
                    if len(price_info) >= 4:
                        coin_data['김프액'] = price_info[3]
                    
                    coins_data.append(coin_data)
                    count += 1
                    
                    # 진행 상황 출력
                    msg = f"  {count}. {coin_name} ({coin_symbol})"
                    if '김프율' in coin_data:
                        msg += f" - 김프율: {coin_data['김프율']}"
                    self.progress_update.emit(msg)
                        
                except Exception as e:
                    self.progress_update.emit(f"[WARNING] 행 {idx} 처리 중 오류: {str(e)}")
                    continue
            
            self.progress_update.emit(f"\n[OK] 총 {len(coins_data)}개 코인 데이터 수집 완료!")
            self.data_scraped.emit(coins_data)
            
        except Exception as e:
            self.error.emit(f"크롤링 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()
            
        finally:
            if self.driver:
                self.driver.quit()
                self.progress_update.emit("[CLOSE] 브라우저 종료")
            self.finished.emit()


class KimpgaScraperGUI(QMainWindow):
    """김프가 스크래퍼 GUI 메인 윈도우"""
    
    def __init__(self):
        super().__init__()
        self.scraper_thread = None
        self.current_data = []
        self.init_ui()
        
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle('김프가(Kimpga) 코인 크롤러 v2.0')
        self.setGeometry(100, 100, 1200, 800)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 메인 레이아웃
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # 타이틀
        title_label = QLabel('🪙 김프가 코인 크롤러')
        title_font = QFont('Arial', 18, QFont.Bold)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50; padding: 10px;")
        main_layout.addWidget(title_label)
        
        # 컨트롤 패널
        control_group = QGroupBox("설정")
        control_layout = QHBoxLayout()
        
        # 코인 개수 설정
        control_layout.addWidget(QLabel('크롤링할 코인 개수:'))
        self.num_coins_spinbox = QSpinBox()
        self.num_coins_spinbox.setMinimum(1)
        self.num_coins_spinbox.setMaximum(100)
        self.num_coins_spinbox.setValue(20)
        self.num_coins_spinbox.setFixedWidth(80)
        control_layout.addWidget(self.num_coins_spinbox)
        
        # Headless 모드 체크박스
        self.headless_checkbox = QCheckBox('백그라운드 모드 (브라우저 숨김)')
        self.headless_checkbox.setChecked(True)
        control_layout.addWidget(self.headless_checkbox)
        
        control_layout.addStretch()
        
        # 시작 버튼
        self.start_button = QPushButton('🚀 크롤링 시작')
        self.start_button.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-size: 14px;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.start_button.clicked.connect(self.start_scraping)
        control_layout.addWidget(self.start_button)
        
        # 중지 버튼
        self.stop_button = QPushButton('⏹ 중지')
        self.stop_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                font-size: 14px;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.stop_button.clicked.connect(self.stop_scraping)
        self.stop_button.setEnabled(False)
        control_layout.addWidget(self.stop_button)
        
        control_group.setLayout(control_layout)
        main_layout.addWidget(control_group)
        
        # 진행 상황 표시
        progress_group = QGroupBox("진행 상황")
        progress_layout = QVBoxLayout()
        
        self.progress_text = QTextEdit()
        self.progress_text.setReadOnly(True)
        self.progress_text.setMaximumHeight(150)
        self.progress_text.setStyleSheet("""
            QTextEdit {
                background-color: #2c3e50;
                color: #ecf0f1;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 11px;
                padding: 5px;
            }
        """)
        progress_layout.addWidget(self.progress_text)
        
        progress_group.setLayout(progress_layout)
        main_layout.addWidget(progress_group)
        
        # 데이터 테이블
        data_group = QGroupBox("수집된 데이터")
        data_layout = QVBoxLayout()
        
        self.data_table = QTableWidget()
        self.data_table.setColumnCount(8)
        self.data_table.setHorizontalHeaderLabels(['순위', '코인명', '심볼', '국내가격', '해외가격', '김프율', '김프액', '수집시간'])
        
        # 테이블 스타일 설정
        self.data_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #f8f9fa;
                gridline-color: #dee2e6;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }
        """)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        data_layout.addWidget(self.data_table)
        
        # 내보내기 버튼
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        
        self.csv_button = QPushButton('📄 CSV로 저장')
        self.csv_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.csv_button.clicked.connect(self.save_to_csv)
        self.csv_button.setEnabled(False)
        export_layout.addWidget(self.csv_button)
        
        self.excel_button = QPushButton('📊 Excel로 저장')
        self.excel_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.excel_button.clicked.connect(self.save_to_excel)
        self.excel_button.setEnabled(False)
        export_layout.addWidget(self.excel_button)
        
        # coin_results.xlsx로 저장 버튼
        self.coin_results_button = QPushButton('💾 coin_results.xlsx로 저장')
        self.coin_results_button.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.coin_results_button.clicked.connect(self.save_to_coin_results)
        self.coin_results_button.setEnabled(False)
        export_layout.addWidget(self.coin_results_button)
        
        data_layout.addLayout(export_layout)
        data_group.setLayout(data_layout)
        main_layout.addWidget(data_group)
        
        # 상태바
        self.statusBar().showMessage('준비')
        
    def start_scraping(self):
        """스크래핑 시작"""
        num_coins = self.num_coins_spinbox.value()
        headless = self.headless_checkbox.isChecked()
        
        # UI 상태 변경
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.csv_button.setEnabled(False)
        self.excel_button.setEnabled(False)
        self.coin_results_button.setEnabled(False)
        self.progress_text.clear()
        self.data_table.setRowCount(0)
        self.current_data = []
        
        # 스크래퍼 스레드 시작
        self.scraper_thread = ScraperThread(num_coins=num_coins, headless=headless)
        self.scraper_thread.progress_update.connect(self.update_progress)
        self.scraper_thread.data_scraped.connect(self.display_data)
        self.scraper_thread.finished.connect(self.scraping_finished)
        self.scraper_thread.error.connect(self.show_error)
        self.scraper_thread.start()
        
        self.statusBar().showMessage('크롤링 진행 중...')
        
    def stop_scraping(self):
        """스크래핑 중지"""
        if self.scraper_thread:
            self.scraper_thread.stop()
            self.update_progress("[STOP] 사용자가 중지했습니다.")
            self.statusBar().showMessage('중지됨')
        
    def update_progress(self, message):
        """진행 상황 업데이트"""
        self.progress_text.append(message)
        # 자동 스크롤
        self.progress_text.verticalScrollBar().setValue(
            self.progress_text.verticalScrollBar().maximum()
        )
        
    def display_data(self, data):
        """데이터 테이블에 표시"""
        self.current_data = data
        self.data_table.setRowCount(len(data))
        
        for row_idx, coin in enumerate(data):
            # 순위
            self.data_table.setItem(row_idx, 0, QTableWidgetItem(str(coin.get('순위', ''))))
            # 코인명
            self.data_table.setItem(row_idx, 1, QTableWidgetItem(coin.get('코인명', '')))
            # 심볼
            self.data_table.setItem(row_idx, 2, QTableWidgetItem(coin.get('심볼', '')))
            # 국내가격
            self.data_table.setItem(row_idx, 3, QTableWidgetItem(coin.get('국내가격', '')))
            # 해외가격
            self.data_table.setItem(row_idx, 4, QTableWidgetItem(coin.get('해외가격', '')))
            # 김프율
            premium_item = QTableWidgetItem(coin.get('김프율', ''))
            # 김프율에 따라 색상 변경
            if '김프율' in coin:
                premium_text = coin['김프율']
                if '-' in premium_text:
                    premium_item.setForeground(QColor('#e74c3c'))  # 빨간색
                else:
                    premium_item.setForeground(QColor('#27ae60'))  # 녹색
            self.data_table.setItem(row_idx, 5, premium_item)
            # 김프액
            self.data_table.setItem(row_idx, 6, QTableWidgetItem(coin.get('김프액', '')))
            # 수집시간
            self.data_table.setItem(row_idx, 7, QTableWidgetItem(coin.get('수집시간', '')))
        
        # 내보내기 버튼 활성화
        self.csv_button.setEnabled(True)
        self.excel_button.setEnabled(True)
        self.coin_results_button.setEnabled(True)
        
    def scraping_finished(self):
        """스크래핑 완료"""
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.statusBar().showMessage('완료')
        
    def show_error(self, error_message):
        """에러 메시지 표시"""
        QMessageBox.critical(self, '오류', error_message)
        self.update_progress(f"[ERROR] {error_message}")
        
    def save_to_csv(self):
        """CSV 파일로 저장"""
        if not self.current_data:
            QMessageBox.warning(self, '경고', '저장할 데이터가 없습니다.')
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, 
            'CSV 파일 저장', 
            f'kimpga_top{len(self.current_data)}_coins_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv',
            'CSV Files (*.csv)'
        )
        
        if filename:
            try:
                df = pd.DataFrame(self.current_data)
                df.to_csv(filename, index=False, encoding='utf-8-sig')
                QMessageBox.information(self, '성공', f'데이터가 저장되었습니다:\n{filename}')
                self.update_progress(f"[SAVE] CSV 파일 저장: {filename}")
            except Exception as e:
                QMessageBox.critical(self, '오류', f'저장 중 오류 발생:\n{str(e)}')
                
    def save_to_excel(self):
        """Excel 파일로 저장"""
        if not self.current_data:
            QMessageBox.warning(self, '경고', '저장할 데이터가 없습니다.')
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, 
            'Excel 파일 저장', 
            f'kimpga_top{len(self.current_data)}_coins_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'Excel Files (*.xlsx)'
        )
        
        if filename:
            try:
                df = pd.DataFrame(self.current_data)
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='김프가 Top Coins')
                    
                    # 워크시트 가져오기
                    worksheet = writer.sheets['김프가 Top Coins']
                    
                    # 열 너비 자동 조정
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                QMessageBox.information(self, '성공', f'데이터가 저장되었습니다:\n{filename}')
                self.update_progress(f"[SAVE] Excel 파일 저장: {filename}")
            except Exception as e:
                QMessageBox.critical(self, '오류', f'저장 중 오류 발생:\n{str(e)}')
    
    def save_to_coin_results(self):
        """coin_results.xlsx 파일로 저장 (openpyxl 사용)"""
        if not self.current_data:
            QMessageBox.warning(self, '경고', '저장할 데이터가 없습니다.')
            return
        
        filename = 'coin_results.xlsx'
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = '김프가 코인 데이터'
            
            # 헤더 정의
            headers = ['순위', '코인명', '심볼', '국내가격', '해외가격', '김프율', '김프액', '수집시간']
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 헤더 작성
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 데이터 작성
            for row_num, coin in enumerate(self.current_data, 2):
                # 순위
                cell = ws.cell(row=row_num, column=1)
                cell.value = coin.get('순위', '')
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                
                # 코인명
                cell = ws.cell(row=row_num, column=2)
                cell.value = coin.get('코인명', '')
                cell.font = Font(bold=True)
                cell.border = border
                
                # 심볼
                cell = ws.cell(row=row_num, column=3)
                cell.value = coin.get('심볼', '')
                cell.border = border
                
                # 국내가격
                cell = ws.cell(row=row_num, column=4)
                cell.value = coin.get('국내가격', '')
                cell.alignment = Alignment(horizontal="right")
                cell.border = border
                
                # 해외가격
                cell = ws.cell(row=row_num, column=5)
                cell.value = coin.get('해외가격', '')
                cell.alignment = Alignment(horizontal="right")
                cell.border = border
                
                # 김프율 (색상 적용)
                cell = ws.cell(row=row_num, column=6)
                premium_value = coin.get('김프율', '')
                cell.value = premium_value
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                
                # 김프율에 따라 색상 변경
                if premium_value and '-' in str(premium_value):
                    cell.font = Font(color="E74C3C", bold=True)  # 빨간색
                elif premium_value:
                    cell.font = Font(color="27AE60", bold=True)  # 녹색
                
                # 김프액
                cell = ws.cell(row=row_num, column=7)
                cell.value = coin.get('김프액', '')
                cell.alignment = Alignment(horizontal="right")
                cell.border = border
                
                # 수집시간
                cell = ws.cell(row=row_num, column=8)
                cell.value = coin.get('수집시간', '')
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            # 열 너비 자동 조정
            column_widths = {
                'A': 8,   # 순위
                'B': 20,  # 코인명
                'C': 12,  # 심볼
                'D': 18,  # 국내가격
                'E': 18,  # 해외가격
                'F': 12,  # 김프율
                'G': 18,  # 김프액
                'H': 20   # 수집시간
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 행 높이 설정
            ws.row_dimensions[1].height = 25
            
            # 파일 저장
            wb.save(filename)
            
            QMessageBox.information(
                self, 
                '성공', 
                f'데이터가 저장되었습니다!\n\n파일: {filename}\n코인 수: {len(self.current_data)}개'
            )
            self.update_progress(f"[SAVE] coin_results.xlsx 파일 저장 완료 ({len(self.current_data)}개 코인)")
            
        except ImportError:
            QMessageBox.critical(
                self, 
                '오류', 
                'openpyxl 패키지가 설치되어 있지 않습니다.\n\npip install openpyxl'
            )
        except Exception as e:
            QMessageBox.critical(self, '오류', f'저장 중 오류 발생:\n{str(e)}')
            self.update_progress(f"[ERROR] 저장 실패: {str(e)}")


def main():
    """메인 실행 함수"""
    app = QApplication(sys.argv)
    
    # 애플리케이션 스타일 설정
    app.setStyle('Fusion')
    
    # 메인 윈도우 생성 및 표시
    window = KimpgaScraperGUI()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
